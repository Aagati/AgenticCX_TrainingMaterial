"""
Builds two workbooks for Day 5 Pre-Lunch Lab H3 (Retail — CX ROI Model):
  - Retail_ROI_Calculator_TEMPLATE.xlsx  (participant fills in yellow cells)
  - Retail_ROI_Calculator_SOLUTION.xlsx  (facilitator reference, all formulas filled)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import copy

BLUE = Font(name="Arial", size=11, color="0000FF")
BLACK = Font(name="Arial", size=11, color="000000")
BOLD = Font(name="Arial", size=11, bold=True, color="000000")
HEADER_FONT = Font(name="Arial", size=12, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=16, bold=True, color="16233F")
SUBTLE_FONT = Font(name="Arial", size=10, italic=True, color="5B6B82")
HEADER_FILL = PatternFill("solid", fgColor="16233F")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # yellow: fill-me-in
GIVEN_FILL = PatternFill("solid", fgColor="E8F1F3")   # light teal: given data
THIN = Side(style="thin", color="DCE3EA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT = "0.0%"
CUR = "$#,##0.00"
CUR0 = "$#,##0"


def style_header(ws, row, cols, text=None):
    if text:
        ws.cell(row=row, column=cols[0], value=text)
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER


def label_cell(ws, row, col, text, bold=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD if bold else BLACK
    return cell


def given(ws, row, col, value, fmt=None, comment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BLUE
    cell.fill = GIVEN_FILL
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if comment:
        cell.comment = Comment(comment, "Facilitator")
    return cell


def formula_cell(ws, row, col, formula, fmt=None, fill_input=False):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = BLACK
    cell.border = BORDER
    cell.fill = INPUT_FILL if fill_input else PatternFill(fill_type=None)
    if fmt:
        cell.number_format = fmt
    return cell


def build(is_template: bool, out_path: str):
    wb = Workbook()

    # ---------------- Instructions ----------------
    ws = wb.active
    ws.title = "Instructions"
    ws.column_dimensions["A"].width = 100
    ws["A1"] = "Day 5 · Pre-Lunch Lab H3 (Retail) — CX ROI Model"
    ws["A1"].font = TITLE_FONT
    lines = [
        "",
        "OBJECTIVE",
        "Compute a CX ROI model for a retail support agent (order status, returns, delivery "
        "issues) using resolution rate, average handle time (AHT) and CSAT — the numbers a "
        "business sponsor would ask for.",
        "",
        "HOW THIS WORKBOOK IS ORGANISED",
        "  - 'Inputs' — the given baseline (pre-agent) and agent-performance data. Blue text on "
        "a light-teal fill = given data, do not change it.",
        "  - 'ROI Model' — where you build the calculation. Cells shaded yellow are for YOU to "
        "fill in with a formula that references the Inputs sheet — never hardcode a number.",
        "  - 'Summary' — the one-paragraph business case, auto-populated from your formulas.",
        "",
        "RULES",
        "  1. Every calculated cell must be a formula, not a typed-in number — so the model "
        "updates if the inputs change.",
        "  2. Reference cells on the Inputs sheet by name (e.g. =Inputs!B5), don't retype values.",
        "  3. The business case is only valid if CSAT stays at or above the floor — check the "
        "PASS/FAIL cell before you trust the savings number.",
        "",
        "WHAT 'DONE' LOOKS LIKE",
        "A completed 'ROI Model' sheet with no yellow cells left blank, and a 'Summary' sheet "
        "that reads as a short business case a retail operations sponsor could act on.",
        "",
        "STRETCH GOAL (optional)",
        "Add a sensitivity check: what resolution rate would be needed to break even if the "
        "agent's build & run cost doubled?",
    ]
    r = 2
    for line in lines:
        cell = ws.cell(row=r, column=1, value=line)
        if line.isupper() or (line and line == line.upper() and not line.startswith(" ")):
            cell.font = BOLD
        else:
            cell.font = BLACK
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # ---------------- Inputs ----------------
    ws2 = wb.create_sheet("Inputs")
    for col, width in zip("ABCDE", [34, 16, 16, 16, 40]):
        ws2.column_dimensions[col].width = width

    ws2["A1"] = "Given Data — Retail CX ROI Model"
    ws2["A1"].font = TITLE_FONT
    ws2["A2"] = "Baseline scenario: order-status, returns and delivery-issue contacts."
    ws2["A2"].font = SUBTLE_FONT

    style_header(ws2, 4, [1, 2], "Metric")
    ws2.cell(row=4, column=2, value="Value")
    style_header(ws2, 4, [1, 2])

    rows = [
        ("Monthly contact volume (all channels)", 20000, "#,##0", None),
        ("Human-handled AHT (minutes)", 9.5, "0.0", "Average handle time for a human agent handling this contact type."),
        ("Fully-loaded human cost per hour", 28, CUR0, "Wages + overhead for a support agent, per hour."),
        ("Agent resolution rate", 0.72, PCT, "Share of contacts the CX agent resolves end-to-end, no human hand-off."),
        ("Agent-handled AHT (minutes)", 3.2, "0.0", "Average handle time for contacts the agent resolves itself."),
        ("Agent cost per contact (tokens + infra)", 0.35, CUR, "Blended compute + tooling cost per contact the agent handles."),
        ("Baseline CSAT (human-only)", 4.2, "0.00", "Out of 5."),
        ("CSAT with agent live", 4.1, "0.00", "Out of 5 — measured post-launch."),
        ("CSAT floor (minimum acceptable)", 4.0, "0.00", "Business-agreed quality guardrail — see Day 5, Topic 4."),
    ]
    r = 5
    for label, val, fmt, note in rows:
        label_cell(ws2, r, 1, label)
        given(ws2, r, 2, val, fmt, note)
        r += 1

    # ---------------- ROI Model ----------------
    ws3 = wb.create_sheet("ROI Model")
    for col, width in zip("ABCDE", [42, 18, 40, 4, 4]):
        ws3.column_dimensions[col].width = width
    ws3["A1"] = "ROI Model"
    ws3["A1"].font = TITLE_FONT
    ws3["A2"] = "Yellow cells = build the formula yourself, referencing the Inputs sheet."
    ws3["A2"].font = SUBTLE_FONT

    style_header(ws3, 4, [1, 2, 3], "Calculation")
    ws3.cell(row=4, column=2, value="Result")
    ws3.cell(row=4, column=3, value="Formula guidance (delete once filled in)")

    def row_entry(row, label, formula_or_blank, fmt, guidance, is_input_row=True):
        label_cell(ws3, row, 1, label)
        if is_template and is_input_row:
            formula_cell(ws3, row, 2, "", fmt, fill_input=True)
        else:
            formula_cell(ws3, row, 2, formula_or_blank, fmt, fill_input=False)
        g = ws3.cell(row=row, column=3, value=guidance if is_template else "")
        g.font = SUBTLE_FONT
        g.alignment = Alignment(wrap_text=True)

    row_entry(5, "Human cost per contact",
              "=(Inputs!B6/60)*Inputs!B7", CUR,
              "(Human AHT in hours) × (human cost per hour). Hint: Inputs!B6 is in minutes.")
    row_entry(6, "Contacts resolved by agent",
              "=Inputs!B5*Inputs!B8", "#,##0",
              "Monthly volume × agent resolution rate.")
    row_entry(7, "Contacts still needing a human",
              "=Inputs!B5-B6", "#,##0",
              "Monthly volume minus contacts resolved by the agent.")
    row_entry(8, "Cost if 100% human-handled (baseline)",
              "=Inputs!B5*B5", CUR0,
              "Monthly volume × human cost per contact.")
    row_entry(9, "Blended cost with agent live",
              "=(B6*Inputs!B10)+(B7*B5)", CUR0,
              "(Agent-resolved contacts × agent cost per contact) + (human-needed contacts × human cost per contact). Note: Inputs!B10 is 'Agent cost per contact' — not B9, which is AHT.")
    row_entry(10, "Monthly savings",
              "=B8-B9", CUR0,
              "Baseline cost minus blended cost with the agent live.")
    row_entry(11, "Blended cost per contact (post-agent)",
              "=B9/Inputs!B5", CUR,
              "Blended cost with agent ÷ monthly volume.")
    row_entry(12, "CSAT floor check",
              '=IF(Inputs!B12>=Inputs!B13,"PASS","FAIL")', "General",
              'IF the live CSAT (Inputs!B12) is >= the floor (Inputs!B13), "PASS", else "FAIL".')
    row_entry(13, "Annualised savings",
              "=B10*12", CUR0,
              "Monthly savings × 12. Only trust this if row 12 = PASS.")

    for row in range(5, 14):
        ws3.cell(row=row, column=2).border = BORDER

    # ---------------- Summary ----------------
    ws4 = wb.create_sheet("Summary")
    ws4.column_dimensions["A"].width = 100
    ws4["A1"] = "Business Case Summary"
    ws4["A1"].font = TITLE_FONT
    ws4["A3"] = "Auto-generated from the ROI Model sheet — do not edit by hand."
    ws4["A3"].font = SUBTLE_FONT
    if is_template:
        summary_formula = (
            '="At current volume, the CX agent resolves "&TEXT(\'Inputs\'!B8,"0%")&'
            '" of contacts, saving an estimated $"&TEXT(\'ROI Model\'!B10,"#,##0")&'
            '" per month ($"&TEXT(\'ROI Model\'!B13,"#,##0")&" annualised), while CSAT is "&'
            '\'ROI Model\'!B12&" against the quality floor."'
        )
    else:
        summary_formula = (
            '="At current volume, the CX agent resolves "&TEXT(Inputs!B8,"0%")&'
            '" of contacts, saving an estimated $"&TEXT(\'ROI Model\'!B10,"#,##0")&'
            '" per month ($"&TEXT(\'ROI Model\'!B13,"#,##0")&" annualised), while CSAT is "&'
            '\'ROI Model\'!B12&" against the quality floor."'
        )
    cell = ws4.cell(row=5, column=1, value=summary_formula)
    cell.font = Font(name="Arial", size=12, bold=True, color="16233F")
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws4.row_dimensions[5].height = 60

    wb.save(out_path)


if __name__ == "__main__":
    build(True, "/home/claude/day5/exercises/03_Prelunch_H3_Retail_ROI_Model/Retail_ROI_Calculator_TEMPLATE.xlsx")
    build(False, "/home/claude/day5/exercises/03_Prelunch_H3_Retail_ROI_Model/Retail_ROI_Calculator_SOLUTION.xlsx")
    print("done")
